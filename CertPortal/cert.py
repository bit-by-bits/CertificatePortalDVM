import io
import os
import sys
import re

from openpyxl import load_workbook
from PyPDF2 import PdfFileWriter, PdfFileReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4

from CertificatePortalDVM.settings import MEDIA_ROOT


def generate_certificate(name, college, position, event, path):
    """
    All four arguments are strings.
    """
    packet = io.BytesIO()
    # create a new PDF with Reportlab
    can = canvas.Canvas(packet, pagesize=A4)

    font_name = 'Helvetica-Bold'
    font_size = 14
    can.setFont(font_name, font_size)

    name_width = stringWidth(name, font_name, font_size)
    college_width = stringWidth(college, font_name, font_size)
    position_width = stringWidth(position, font_name, font_size)
    event_width = stringWidth(event, font_name, font_size)

    if name_width >= 225:
        can.drawString(356, 288, name)
    else:
        can.drawString(468 - name_width / 2, 288, name)

    if college_width >= 350:
        can.drawString(273, 260, college)
    else:
        can.drawString(445 - college_width / 2, 260, college)

    can.drawString(445 - position_width / 2, 232, position)
    can.drawString(492 - event_width / 2, 204, event)
    can.save()

    # move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)
    # read your existing PDF
    # existing_pdf = PdfFileReader(open("certi.pdf", "rb"), strict=False)
    existing_pdf_loc = os.path.dirname(os.path.abspath(__file__))
    existing_pdf = PdfFileReader(open(os.path.join(existing_pdf_loc, "certi.pdf"), "rb"), strict=False)
    # the script was not able to find the location of the certi.pdf hence this line was modified

    output = PdfFileWriter()
    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page.mergePage(new_pdf.getPage(0))
    output.addPage(page)
    # finally, write "output" to a real file
    file_name = name + '.pdf'
    outputStream = open(path + '/' + file_name, "wb")
    output.write(outputStream)
    outputStream.close()


NAME = 1
COLLEGE = 2
POSITION = 3
EVENT = 4

# # parent_dir = os.path.dirname(os.path.abspath(__file__))
# parent_dir = MEDIA_ROOT
# # parent_dir was changed as the Excel file will be stored in the media directory
# loc = (os.path.join(parent_dir, './Winners.xlsx'))
#
# wb_read = load_workbook(filename=loc)
# sheet = wb_read.active


def generate(sheet, time_stamp):
    count = 0
    max_row = 0
    for row in sheet:
        if not all([cell.value == None for cell in row]):
            max_row += 1

    for row in range(2, max_row + 1):
        name = str(sheet.cell(row=row, column=NAME).value)
        college = str(sheet.cell(row=row, column=COLLEGE).value)
        position = str(sheet.cell(row=row, column=POSITION).value)
        event = str(sheet.cell(row=row, column=EVENT).value)

        # Adding regex could fix the problem of special character causing issues with directory and path names
        # potential solution :

        pattern = r'[^A-Za-z0-9]+'
        name = re.sub(pattern, '', name)
        college = re.sub(pattern, '', name)
        position = re.sub(pattern, '', position)
        event = re.sub(pattern, '', event)


        try:
            path = os.path.join(MEDIA_ROOT, f'./Certificates{time_stamp}/{event}')
            os.makedirs(path)
        except FileExistsError:
            pass
        if college == 'None':
            college = '-'

        generate_certificate(name, college, position, event, path)
        count += 1
        # print(str(count) + ": Certificate generated for " + name + ' - ' + college)

# generate()

# generate_certificate(
#     "Jay Panchal", 
#     "Birla Institute of Technology & Science (BITS), Pilani",
#     "2nd",
#     "CAD Competition", "./")
